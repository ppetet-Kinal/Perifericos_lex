import os
import pandas as pd
import tkinter as tk
from tkinter import messagebox
from tkinter import ttk
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


root = tk.Tk()
root.title("Reporte de Periféricos")
root.geometry("500x525")

ruta_tema_azure = r"C:\Users\Soporte 2\Documents\Phyton\Pruebas Python\Azure-ttk-theme-main\azure.tcl"
try:
    if os.path.exists(ruta_tema_azure):
        root.tk.call("source", ruta_tema_azure)
        root.tk.call("set_theme", "dark")
        print("Tema Azure aplicado con éxito.")
    else:
        raise FileNotFoundError("El archivo del tema Azure no se encontró.")
except FileNotFoundError as e:
    print(e)
    print("Usando el tema predeterminado.")
except Exception as e:
    print(f"Se produjo un error al cargar el tema Azure: {e}")
    print("Usando el tema predeterminado.")

ruta_icono = r"C:\Users\Soporte 2\Documents\Phyton\Control Perofericos\Logo.ico"
if os.path.exists(ruta_icono):
    root.iconbitmap(ruta_icono)


def enviar_reporte():
    nombre = entrada_nombre.get()
    codigo_cartera = entrada_codigo.get()
    nivel_prioridad = combobox_nivel.get()
    reporte = cuadro_texto.get("1.0", "end-1c")
    
   
    if not nombre or not codigo_cartera or not nivel_prioridad or not reporte:
        messagebox.showerror("Error", "Por favor, complete todos los campos.")
        return

 
    datos = {
        "Nombre": nombre,
        "Código Cartera": codigo_cartera,
        "Nivel de Prioridad": nivel_prioridad,
        "Reporte de Periféricos": reporte
    }

  
    archivo_excel = r"X:\RED\Backup\Datos de sistema\Datos Excel\Reporte de Perifericos\Reporte de Perifericos.xlsx"
    carpeta = os.path.dirname(archivo_excel)

    
    if not os.access(carpeta, os.W_OK):
        try:
            if not os.path.exists(carpeta):
                os.makedirs(carpeta)
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo acceder o crear la carpeta:\n{e}")
            return

    try:
       
        if os.path.exists(archivo_excel):
            df = pd.read_excel(archivo_excel)
            df = pd.concat([df, pd.DataFrame([datos])], ignore_index=True)
        else:
            df = pd.DataFrame([datos])

        
        df = df.sort_values(by="Nivel de Prioridad", key=lambda x: x.map({"Alta": 0, "Media": 1, "Baja": 2}))

       
        df.to_excel(archivo_excel, index=False)

       
        wb = load_workbook(archivo_excel)
        ws = wb.active
        colores = {
            "Alta": "FF4500",  # Rojo
            "Media": "FFD700",  # Amarillo
            "Baja": "32CD32"   # Verde
        }

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=3, max_col=3):
            for cell in row:
                if cell.value in colores:
                    cell.fill = PatternFill(start_color=colores[cell.value], end_color=colores[cell.value], fill_type="solid")

        wb.save(archivo_excel)
        wb.close()

        messagebox.showinfo("Éxito", "Reporte enviado correctamente.")
    except PermissionError:
        messagebox.showerror("Error de permisos", f"No se pudo acceder al archivo:\n{archivo_excel}\nVerifique permisos o si el archivo está abierto.")
    except Exception as e:
        messagebox.showerror("Error inesperado", f"Se produjo un error:\n{e}")

label_titulo = tk.Label(root, text="REPORTE DE ESTADO DE PERIFÉRICOS", font=("Adobe Garamond Pro", 18))
label_titulo.pack(pady=10)

label_nombre = tk.Label(root, text="Nombre:")
label_nombre.pack(fill="x", padx=10, pady=5)
entrada_nombre = tk.Entry(root, width=67, background='gray', fg='white')
entrada_nombre.pack(pady=5)

label_codigo_cartera = tk.Label(root, text="Código Cartera:")
label_codigo_cartera.pack(fill="x", padx=10, pady=5)
entrada_codigo = tk.Entry(root, width=67, background='gray', fg='white')
entrada_codigo.pack(pady=5)

label_nivel = tk.Label(root, text="Nivel de Prioridad:")
label_nivel.pack(fill="x", padx=10, pady=5)
niveles_prioridad = ["Baja", "Media", "Alta"]
combobox_nivel = ttk.Combobox(root, values=niveles_prioridad, state="readonly", width=37)
combobox_nivel.pack(fill="x", padx=10, pady=5)

label_reporte = tk.Label(root, text="Reporte de Periféricos:")
label_reporte.pack(fill="x", padx=10, pady=5)
cuadro_texto = tk.Text(root, height=10, width=65)
cuadro_texto.pack(pady=5)

boton_enviar = tk.Button(root, text="Enviar", command=enviar_reporte)
boton_enviar.pack(side='left', padx=15, pady=5)

boton_cerrar = tk.Button(root, text="Cerrar", command=root.destroy)
boton_cerrar.pack(side='right', padx=15, pady=5)


root.mainloop()
